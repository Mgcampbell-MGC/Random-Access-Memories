#!/usr/bin/env python3
"""Sutphin — Portaria 140/2026 curtailment call list. Data tabs."""
import json, collections, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCR = "/tmp/claude-0/-home-user-Random-Access-Memories/5dfc65b1-c39e-5910-843d-b38c7c05b704/scratchpad"
FONT, NAVY, RULE = "Arial", "1F3864", "BFBFBF"
HDRFILL = PatternFill("solid", fgColor=NAVY)
SUBFILL = PatternFill("solid", fgColor="D9E2F3")
YEL     = PatternFill("solid", fgColor="FFFF00")
AMBER   = PatternFill("solid", fgColor="FFF2CC")
BLUE, BLACK, GREEN = "0000FF", "000000", "008000"
THIN = Side(style="thin", color=RULE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BRL = 'R$ #,##0;(R$ #,##0);-'
NUM = '#,##0;(#,##0);-'
PCT = '0.0%'

def hdr(ws, row, labels, widths=None):
    for i, lab in enumerate(labels, 1):
        c = ws.cell(row=row, column=i, value=lab)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = HDRFILL; c.border = BOX
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    ws.row_dimensions[row].height = 32
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")

def put(ws, r, c, v, *, bold=False, size=9, color=BLACK, wrap=True, fill=None,
        fmt=None, italic=False, align="left"):
    x = ws.cell(row=r, column=c, value=v)
    x.font = Font(name=FONT, size=size, bold=bold, color=color, italic=italic)
    x.alignment = Alignment(wrap_text=wrap, vertical="top", horizontal=align)
    x.border = BOX
    if fill: x.fill = fill
    if fmt: x.number_format = fmt
    return x

CX = json.load(open(f"{SCR}/complexes.json"))
TOT = sum(c["elig"] for c in CX)

# ---------------------------------------------------------------- Assumptions
ASSUM = [
    ("INPUTS — edit the yellow cells; every estimate on the other tabs recalculates", None, None, None),
    ("Compensation pool, Portaria 140 window", 3300, "R$ m",
     "MME's own estimate as reported 21–23 Jul 2026. Volt Robotics says R$2,700m; ABEEólica R$3,300m; Folhapress reports MME at R$3,490m. NO PRIMARY DOCUMENT STATES ANY FIGURE — the Portaria itself sets no pool."),
    ("Eligible curtailment in the window", 25326, "GWh",
     "COMPUTED, not sourced. Built from ONS half-hourly constrained-off open data for 1 Sep 2023 – 25 Nov 2025, counting only REL (indisponibilidade externa) and CNF (confiabilidade elétrica). Excludes ENE (sobreoferta), which the Portaria does not pay for."),
    ("Implied compensation price", None, "R$/MWh",
     "= pool / eligible volume. Sits between the PLD floor (R$57.31 in 2026) and ACR contract prices, so the pool estimates are internally coherent."),
    ("IPCA uplift, event date to payment", 0.22, "%",
     "EDITABLE. Art. 6º §1º indexes from each event date to actual payment. Events run Sep 2023–Nov 2025; payment late 2027–H1 2028. A rough 22% assumes ~4.5%/yr over ~4.4 years average. Replace with a proper monthly IPCA build before pricing anything."),
    ("Sutphin fee on an advisory mandate", 0.015, "%",
     "EDITABLE. Illustrative only. No mandate has been signed and no fee has been discussed with anyone."),
    ("", None, None, None),
    ("WHAT THIS MODEL DELIBERATELY DOES NOT DO", None, None, None),
    ("It does not split cash from book entry", None, None,
     "This is the single most commercially important variable and IT CANNOT BE DERIVED FROM PUBLIC DATA. Cláusula Quinta pays the CCEAR-D/CER-committed tranche by cancelling the generator's own delivery-shortfall debt — no cash moves. Only the uncommitted tranche is paid in cash at PLD. The contract position of each plant sits in CCEE data behind a login. ONLY THE GENERATOR KNOWS. That is why it is the first question on the call."),
    ("It does not net against what they owe CCEE", None, None,
     "Sector-wide, ~R$3.3bn of credit is netted against ~R$6bn generators owe CCEE (ABEEólica). At sector level the credit does not cover the debt. Per company the ratio varies enormously and is not public."),
    ("It does not price the waiver", None, None,
     "Signing waives claims on ALL curtailment to 25 Nov 2025 including sobreoferta — 23,213 GWh across this population that is never compensated. Whether that claim is worth anything depends on the pending Congressional veto override (see Mechanics tab)."),
    ("It does not value the garantia física BENEFIT", None, None,
     "Art. 10 counts cut volumes as generation verificada for the annual garantia física revision, 'passíveis ou não de compensação'. This PROTECTS the plant — it stops the physical guarantee being revised down for output the plant was ordered not to produce, and it covers sobreoferta cuts too, which are never paid in cash. MME's methodology is still undefined so it cannot be quantified, but the direction is favourable and it belongs on the POSITIVE side of the sign/do-not-sign decision. An earlier version of this model had it inverted."),
    ("It uses one price for everyone", None, None,
     "In reality compensation is paid at the PLD of the plant's submarket in the specific hours cut. In the Nordeste those are disproportionately floor hours: the 10h–15h block averaged R$134/MWh in 2026 YTD with 56% of hours at the R$57.31 floor. A plant cut mostly at midday recovers far less than this model implies."),
]

def tab_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    title(ws, "Assumptions and method",
          "Yellow cells are inputs. Everything else is a formula. Read the second block before quoting any number.")
    r = 4
    hdr(ws, r, ["Input", "Value", "Unit", "Basis, and how much to trust it"], [46, 14, 10, 96]); r += 1
    ref = {}
    for lab, val, unit, note in ASSUM:
        if val is None and unit is None:
            put(ws, r, 1, lab, bold=True, size=10, color=NAVY, fill=SUBFILL)
            for c in (2, 3, 4): put(ws, r, c, "", fill=SUBFILL)
        else:
            put(ws, r, 1, lab, bold=True)
            if lab.startswith("Implied"):
                put(ws, r, 2, "=B6/B7*1000", fill=AMBER, fmt='#,##0.0', color=BLACK)
                ref["price"] = f"$B${r}"
            elif val is not None:
                fmt = PCT if unit == "%" else NUM
                put(ws, r, 2, val, fill=YEL, fmt=fmt, color=BLUE)
                key = {"Compensation pool, Portaria 140 window": "pool",
                       "Eligible curtailment in the window": "vol",
                       "IPCA uplift, event date to payment": "ipca",
                       "Sutphin fee on an advisory mandate": "fee"}.get(lab)
                if key: ref[key] = f"$B${r}"
            put(ws, r, 3, unit or "")
            put(ws, r, 4, note or "")
        r += 1
    ws.freeze_panes = "A5"
    return ref

# ---------------------------------------------------------------- Call list
WHY = {
 'Casa dos Ventos': "Largest single holder of eligible claim in the country — and privately held, so no analyst has done this work for them. Rio do Vento and Serra do Mel clusters. Ask what share of the cut volume sat outside CCEAR-D/CER.",
 'Engie Brasil Energia': "Named its own worst assets in its 4T25 release (Santo Agostinho 36%, Trairi 31%). Discloses curtailment only in percentages — has never put an R$ figure on it publicly. Fitch estimates R$150m of 2025 EBITDA impact.",
 'Enel Green Power Brasil': "Lagoa dos Ventos is the largest wind complex in South America. Foreign parent, decision may sit in Rome — find out who signs in Brazil before 10 August.",
 'Echoenergia (Equatorial)': "ALREADY LITIGATING — two of the twelve cases (1068423-72 and 1068397-74.2024.4.01.3400). CEO Liu Aquino has said publicly some plants hit ~45% restriction, 'patamar que torna a operação insustentável'. They must actively choose to abandon those suits.",
 'Serena Energia': "Fitch puts 2025 EBITDA impact at R$200m, 11% of total — the highest proportional hit of any rated name. Listed, so the decision is disclosable and time-pressured.",
 'Voltalia Brasil': "THE ONLY COMPANY THAT HAS BOOKED REAL MONEY — R$175m of curtailment ressarcimento recognised in Q2 2026 results. They understand the mechanism better than anyone. Call them to learn, not to sell.",
 'Elera Renováveis (Brookfield)': "ALREADY LITIGATING — two cases over the Alex complex (0800580-25.2024.4.05.8101 and 1087805-51.2024.4.01.3400). Complexo Alex hit 67.7% curtailment in July 2024. Brookfield capital behind it, so the waiver decision will be taken seriously and centrally.",
 'CPFL Energias Renováveis (State Grid)': "The most transparent discloser in the sector — published R$558m of FY2025 lost revenue and 30.8% curtailment. They have already done the arithmetic; the value you add is on the waiver, not the volume.",
 'Janaúba (Telefônica/Aegea AP)': "Self-production structure. Autoprodução is expressly eligible (Termo Cl. 1ª Subcl. 1ª) and its volume is NOT committed to CCEAR-D/CER — so under Cláusula Quinta V it is valued at PLD and PAID IN CASH. This is the profile most likely to produce a real receivable.",
 'Jaíba V (Sendas/Assaí AP)': "Self-production. Same logic as Janaúba — uncommitted volume means cash rather than a book entry. The counterparty is a retailer, not a power company, so nobody in-house owns this problem.",
 'Belmonte (Mateus Supermercados AP)': "Self-production for a supermarket group. Almost certainly nobody internally is tracking a 10 August federal energy deadline.",
 'Laranjeiras (Cargill AP)': "Self-production for Cargill. Foreign corporate parent, no Brazilian energy-regulatory team of its own.",
}
ASK = ("1) What share of your cut volume was committed to CCEAR-D or CER? "
       "2) What do you owe CCEE in delivery-shortfall ressarcimentos? "
       "3) What is your sobreoferta volume over the window, and what have you assumed it is worth? "
       "4) Are you in any of the twelve suits, or in the ABEEólica/ABSOLAR collective action?")

def tab_calllist(wb, ref):
    by = collections.defaultdict(lambda: [0., 0., 0., [], set(), ''])
    for c in CX:
        b = by[c["group"]]
        b[0] += c["elig"]; b[1] += c["ene"]; b[2] += c["mw"]
        b[3].append(c["complex"]); b[4].add(c["uf"]); b[5] = c["conf"]
    rows = sorted(by.items(), key=lambda x: -x[1][0])
    rows = [r for r in rows if r[0] != '— unmapped —']

    ws = wb.create_sheet("Call List", 0)
    title(ws, "Call list — Portaria 140/2026, ranked by eligible claim",
          "Eligible GWh is COMPUTED from ONS half-hourly constrained-off data for 1 Sep 2023 – 25 Nov 2025, "
          "REL + CNF only. It is not a company disclosure. Ownership is a join through ANEEL SIGA — check the confidence column.")
    r = 4
    hdr(ws, r, ["#", "Group / holder", "Conf.", "States", "Complexes",
                "Eligible GWh", "% of pool", "Est. gross claim (R$m)", "Indexed to payment (R$m)",
                "Sobreoferta GWh (waived, never paid)", "Operating MW", "Why they take the call", "What to ask on call one"],
        [4, 34, 6, 12, 8, 11, 8, 12, 12, 13, 10, 62, 54]); r += 1
    first = r
    for i, (name, (e, ene, mw, cx, ufs, conf)) in enumerate(rows[:40], 1):
        put(ws, r, 1, i, align="center")
        put(ws, r, 2, name, bold=True)
        put(ws, r, 3, conf, align="center",
            color={"H": "008000", "M": "BF8F00", "L": "C00000"}.get(conf, BLACK), bold=True)
        put(ws, r, 4, ",".join(sorted(ufs)))
        put(ws, r, 5, len(cx), fmt=NUM, align="right")
        put(ws, r, 6, round(e, 1), fmt='#,##0', align="right")
        put(ws, r, 7, f"=F{r}/Assumptions!{ref['vol']}", fmt=PCT, align="right")
        put(ws, r, 8, f"=F{r}*1000*Assumptions!{ref['price']}/1000000", fmt=NUM, align="right", bold=True)
        put(ws, r, 9, f"=H{r}*(1+Assumptions!{ref['ipca']})", fmt=NUM, align="right")
        put(ws, r, 10, round(ene, 1), fmt='#,##0', align="right", color="C00000")
        put(ws, r, 11, round(mw, 0), fmt=NUM, align="right")
        put(ws, r, 12, WHY.get(name, "Mapped from SIGA ownership. Confirm the group before calling — the SPE names are in the Complexes tab."))
        put(ws, r, 13, ASK)
        r += 1
    put(ws, r, 2, "TOTAL — named groups above", bold=True, fill=SUBFILL)
    for c in (1, 3, 4, 12, 13): put(ws, r, c, "", fill=SUBFILL)
    put(ws, r, 5, f"=SUM(E{first}:E{r-1})", fmt=NUM, bold=True, fill=SUBFILL, align="right")
    put(ws, r, 6, f"=SUM(F{first}:F{r-1})", fmt='#,##0', bold=True, fill=SUBFILL, align="right")
    put(ws, r, 7, f"=SUM(G{first}:G{r-1})", fmt=PCT, bold=True, fill=SUBFILL, align="right")
    put(ws, r, 8, f"=SUM(H{first}:H{r-1})", fmt=NUM, bold=True, fill=SUBFILL, align="right")
    put(ws, r, 9, f"=SUM(I{first}:I{r-1})", fmt=NUM, bold=True, fill=SUBFILL, align="right")
    put(ws, r, 10, f"=SUM(J{first}:J{r-1})", fmt='#,##0', bold=True, fill=SUBFILL, align="right")
    put(ws, r, 11, f"=SUM(K{first}:K{r-1})", fmt=NUM, bold=True, fill=SUBFILL, align="right")
    r += 2
    put(ws, r, 2, "Confidence: H = corroborated by the company's own disclosure or by a filed lawsuit. "
                  "M = inferred from SPE naming convention in ANEEL SIGA — plausible, not verified. "
                  "L = SPE name only, corporate group not established. Do not put an M or L name in front of GC as fact.",
        italic=True, color="C00000", size=9)
    ws.freeze_panes = "C5"
    return rows

# ---------------------------------------------------------------- Complexes
def tab_complexes(wb, ref):
    ws = wb.create_sheet("Complexes")
    title(ws, "Every complex with eligible curtailment, 1 Sep 2023 – 25 Nov 2025",
          "The underlying evidence for the call list. 267 apuração units, 25,326 GWh of REL+CNF curtailment, "
          "computed from ONS open data. Sobreoferta (ENE) shown separately because the Portaria never pays for it.")
    r = 4
    hdr(ws, r, ["#", "Complex", "UF", "Source", "Group / holder", "Conf.",
                "Eligible GWh (REL+CNF)", "% of pool", "Est. claim (R$m)",
                "Sobreoferta GWh (ENE)", "Eligible share of cuts", "Operating MW", "SPE holders in ANEEL SIGA"],
        [4, 32, 5, 7, 34, 6, 12, 8, 11, 12, 10, 10, 78]); r += 1
    first = r
    for i, c in enumerate(sorted(CX, key=lambda x: -x["elig"]), 1):
        tt = c["elig"] + c["ene"]
        put(ws, r, 1, i, align="center")
        put(ws, r, 2, c["complex"], bold=True)
        put(ws, r, 3, c["uf"], align="center")
        put(ws, r, 4, "Wind" if c["src"] == "EOL" else "Solar", align="center")
        put(ws, r, 5, c["group"])
        put(ws, r, 6, c["conf"], align="center",
            color={"H": "008000", "M": "BF8F00", "L": "C00000"}.get(c["conf"], BLACK), bold=True)
        put(ws, r, 7, round(c["elig"], 1), fmt='#,##0.0', align="right")
        put(ws, r, 8, f"=G{r}/Assumptions!{ref['vol']}", fmt=PCT, align="right")
        put(ws, r, 9, f"=G{r}*1000*Assumptions!{ref['price']}/1000000", fmt='#,##0', align="right", bold=True)
        put(ws, r, 10, round(c["ene"], 1), fmt='#,##0.0', align="right", color="C00000")
        put(ws, r, 11, (c["elig"] / tt) if tt else 0, fmt=PCT, align="right")
        put(ws, r, 12, round(c["mw"], 0), fmt=NUM, align="right")
        put(ws, r, 13, c["spes"])
        r += 1
    put(ws, r, 2, "TOTAL", bold=True, fill=SUBFILL)
    for c in (1, 3, 4, 5, 6, 8, 11, 13): put(ws, r, c, "", fill=SUBFILL)
    put(ws, r, 7, f"=SUM(G{first}:G{r-1})", fmt='#,##0', bold=True, fill=SUBFILL, align="right")
    put(ws, r, 9, f"=SUM(I{first}:I{r-1})", fmt='#,##0', bold=True, fill=SUBFILL, align="right")
    put(ws, r, 10, f"=SUM(J{first}:J{r-1})", fmt='#,##0', bold=True, fill=SUBFILL, align="right")
    put(ws, r, 12, f"=SUM(L{first}:L{r-1})", fmt=NUM, bold=True, fill=SUBFILL, align="right")
    ws.freeze_panes = "C5"

# ---------------------------------------------------------------- Contacts
from contactdata import CONTACTS


def tab_contacts(wb):
    ws = wb.create_sheet("Contacts")
    title(ws, "Named contacts — verified from the public record only",
          "Nobody on this sheet is a guess. Where a name could not be verified it has been left out rather than invented.")
    r = 4
    hdr(ws, r, ["Name", "Role", "Organisation", "Conf.", "What is on the record", "Why this call"],
        [30, 30, 26, 6, 96, 70]); r += 1
    for n, role, org, conf, rec, why in CONTACTS:
        if role is None:
            put(ws, r, 1, n, bold=True, size=10, color=NAVY, fill=SUBFILL)
            for c in range(2, 7): put(ws, r, c, "", fill=SUBFILL)
        else:
            put(ws, r, 1, n, bold=True); put(ws, r, 2, role); put(ws, r, 3, org)
            put(ws, r, 4, conf, align="center", bold=True,
                color={"H": "008000", "M": "BF8F00"}.get(conf, BLACK))
            put(ws, r, 5, rec); put(ws, r, 6, why)
        r += 1
    ws.freeze_panes = "A5"

# ---------------------------------------------------------------- Sources
SOURCES = [
 ("PRIMARY — THE INSTRUMENT", None, None),
 ("Portaria Normativa MME nº 140, de 18 de julho de 2026", "DOU 21/07/2026, Ed. 135-A, Seção 1 Extra A, p.10",
  "https://www.in.gov.br/web/dou/-/portaria-normativa-mme-n-140-de-18-de-julho-de-2026-720529777"),
 ("MME release on the Portaria", "gov.br/mme",
  "https://www.gov.br/mme/pt-br/assuntos/noticias/mme-regulamenta-procedimentos-para-celebracao-de-termo-de-compromisso-sobre-compensacao-por-cortes-de-geracao"),
 ("CELEBRA — the election portal", "Where the manifestação de interesse is filed by 10 Aug 2026", "https://celebra.mme.gov.br/"),
 ("Lei 10.848/2004, art. 1º-B", "Inserted by Lei 15.269/2025 — the statutory basis",
  "https://www.planalto.gov.br/ccivil_03/_ato2004-2006/2004/lei/l10.848.htm"),
 ("Lei 15.269/2025", "Sanctioned 24 Nov 2025, published 25 Nov 2025",
  "https://www.planalto.gov.br/ccivil_03/_ato2023-2026/2025/lei/L15269.htm"),
 ("Veto message VEP-1755-25", "The vetoed art. 1º-A — STILL NOT VOTED BY CONGRESS",
  "https://www.planalto.gov.br/ccivil_03/_ato2023-2026/2025/Msg/Vep/VEP-1755-25.htm"),
 ("ANEEL Voto, Proc. 48500.000231/2026-56", "Diretora Agnes Maria de Aragão da Costa, Jan 2026. Authoritative source for the 46% ENE / 41% CNF / 13% REL split",
  "https://canalsolar.com.br/wp-content/uploads/2026/01/voto.pdf"),
 ("DATA — COMPUTED FOR THIS ANALYSIS", None, None),
 ("ONS constrained-off, wind, per apuração unit", "Half-hourly, Oct 2021 → Jul 2026, CC-BY. The eligible-claim column is built from this.",
  "https://dados.ons.org.br/dataset/restricao_coff_eolica_usi"),
 ("ONS constrained-off, wind, per plant", "1,053 individual plants with CEG — the ownership join",
  "https://dados.ons.org.br/dataset/restricao_coff_eolica_detail"),
 ("ONS constrained-off, solar", "SERIES STARTS APRIL 2024 — there is no public solar curtailment data before that",
  "https://dados.ons.org.br/dataset/restricao_coff_fotovoltaica"),
 ("ANEEL SIGA", "Ownership and capacity by CEG. 98.9% of the CEGs in the ONS detail files matched.",
  "https://dadosabertos.aneel.gov.br/dataset/siga-sistema-de-informacoes-de-geracao-da-aneel"),
 ("CCEE — encargo received per agent and per plant", "Includes an ENCARGO_CONSTRAINED_OFF column, monthly 202311→202605",
  "https://dadosabertos.ccee.org.br/dataset/encargo_recebido_perfil_agente_usina"),
 ("CCEE PLD", "Nordeste monthly average: 2024 R$118.34 · 2025 R$176.60 · 2026 YTD R$226.63. Floor 2026 R$57.31.",
  "https://dadosabertos.ccee.org.br/dataset/pld_horario"),
 ("SIZING — THE COMPETING ESTIMATES", None, None),
 ("MME ~R$3.3bn netted against ~R$6bn owed to CCEE", "Poder360, 22 Jul 2026. NO PRIMARY DOCUMENT STATES ANY POOL FIGURE.",
  "https://www.poder360.com.br/poder-governo/governo-regulamenta-acordo-de-r-33-bi-com-usinas-renovaveis/"),
 ("Volt Robotics — R$2.7bn compensable; classification is a black box", "eixos / MegaWhat, 23 Jul 2026. Found MME's SOSIN formula produces 143% more sobreoferta than ONS's own method.",
  "https://eixos.com.br/politica/volt-robotics-estima-ressarcimento-por-cortes-de-geracao-em-r-27-bilhoes-e-aponta-falta-de-clareza-na-classificacao/"),
 ("Instituto Acende Brasil — R$3.85bn all cuts at PLD, R$2.85bn compensable", "Senate CI hearing, 30 Sep 2025. Best-documented independent methodology.",
  "https://acendebrasil.com.br/wp-content/uploads/2025/09/20250930_AcendeBr_AP_CI_Senado_Curtailment_rev_3-1.pdf"),
 ("Volt Robotics — 2025 balanço: R$6.5bn, 20.6% of available generation", "Jan 2026",
  "https://cenarioenergia.com.br/wp-content/uploads/2026/01/VoltRobotics_AnaliseCurtailment_Balanco-2025_Janeiro-2026.pdf"),
 ("Fitch per-company 2025 EBITDA impact", "Auren R$400m · Serena R$200m · Engie Brasil R$150m. Via NeoFeed, 25 Feb 2026.",
  "https://neofeed.com.br/negocios/brasil-desperdica-20-de-sua-energia-renovavel-e-gera-prejuizos-bilionarios-a-empresas-do-setor-eletrico/"),
 ("LITIGATION", None, None),
 ("The census of twelve suits, with process numbers", "MegaWhat, 31 Oct 2024. None had obtained a liminar.",
  "https://megawhat.energy/geracao/aneel-ja-e-alvo-de-12-acoes-judiciais-por-curtailment/"),
 ("STJ suspends the one TRF1 injunction", "SLS 3.546, Min. Herman Benjamin, Jan 2025",
  "https://www.arandanet.com.br/revista/em/noticia/10214-Aneel-consegue-suspender-liminar-do-constrained-off-no-STJ.html"),
 ("Congress removes the veto override from the agenda", "Canal Solar, 17 Jun 2026 — STILL UNVOTED as at 26 Jul 2026",
  "https://canalsolar.com.br/congresso-analisa-vetos-regras-gd-eolicas-offshore/"),
 ("Industry expects continuing litigation", "eixos, 23 Jul 2026",
  "https://eixos.com.br/energias-renovaveis/portaria-traz-alivio-mas-nao-encerra-judicializacao-sobre-cortes-de-geracao-avaliam-geradores-e-especialistas/"),
 ("COMPANY DISCLOSURE", None, None),
 ("Voltalia books R$175m of curtailment ressarcimento in Q2 2026", "THE FIRST REALISED COMPENSATION FIGURE IN THE MARKET. MegaWhat, 23 Jul 2026.",
  "https://megawhat.uol.com.br/economia-e-politica/resultados/voltalia-contabiliza-r-175-mi-em-ressarcimento-do-curtailment-no-resultado-do-2o-tri/"),
 ("CPFL FY2025 — R$558m lost revenue, 30.8% curtailment", "CPFL 4T25 release. The most transparent disclosure in the sector.",
  "https://ri.cpfl.com.br/"),
 ("Engie 4T25 — curtailment in percentages only, no R$", "Worst assets: Santo Agostinho 36%, Trairi 31%, Campo Largo 17%, Serra do Assuruá 14%",
  "https://www.engie.com.br/wp-content/uploads/2026/02/260225-Release-de-Resultados-4T25.pdf"),
 ("Auren — the only listed generator publishing an explicit R$ curtailment table", "FY2025 gross R$529.5m, R$333.6m net of modulation gains",
  "https://ri.auren.com.br/"),
]

def tab_sources(wb):
    ws = wb.create_sheet("Sources")
    title(ws, "Sources", "Everything in this workbook is traceable to a row below. Where a figure could not be traced to a primary document, the workbook says so.")
    r = 4
    hdr(ws, r, ["Source", "What it gives you, and any caveat", "Link"], [56, 92, 76]); r += 1
    for a, b, c in SOURCES:
        if b is None:
            put(ws, r, 1, a, bold=True, size=10, color=NAVY, fill=SUBFILL)
            put(ws, r, 2, "", fill=SUBFILL); put(ws, r, 3, "", fill=SUBFILL)
        else:
            put(ws, r, 1, a, bold=True); put(ws, r, 2, b)
            x = put(ws, r, 3, c, color="0563C1"); x.hyperlink = c
        r += 1
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------- PF targets
def tab_pf(wb, ref):
    from pfdata import HEADLINE, ROWS, FOOTER
    ws = wb.create_sheet("PF Targets")
    title(ws, "The lending list — Fitch's watch-listed project financings, against our claim data",
          "Two independent sources that agree: Fitch's published Rating Watch Negative list, and our own "
          "eligible-claim computation from raw ONS data. 5,414 GWh — 21.4% of the national pool — sits inside these.")
    r = 4
    for h in HEADLINE:
        put(ws, r, 1, h, bold=(r == 4), color="C00000" if r == 5 else BLACK, size=10 if r < 6 else 9)
        for cc in range(2, 8): put(ws, r, cc, "")
        r += 1
    r += 1
    hdr(ws, r, ["Fitch entity", "Our matched complexes", "UF", "Eligible GWh", "Est. claim (R$m)",
                "Rating and coverage detail", "Sponsor / SPEs", "Named contact", "Why this one"],
        [30, 30, 5, 10, 11, 54, 44, 42, 62]); r += 1
    first = r
    for fe, cxs, uf, gwh, rating, sponsor, contact, why in ROWS:
        put(ws, r, 1, fe, bold=True)
        put(ws, r, 2, cxs)
        put(ws, r, 3, uf, align="center")
        put(ws, r, 4, gwh if gwh else "n/a", fmt=NUM if gwh else None, align="right")
        put(ws, r, 5, f"=IF(ISNUMBER(D{r}),D{r}*1000*Assumptions!{ref['price']}/1000000,\"n/a\")",
            fmt=NUM, align="right", bold=True)
        put(ws, r, 6, rating)
        put(ws, r, 7, sponsor)
        put(ws, r, 8, contact)
        put(ws, r, 9, why)
        r += 1
    put(ws, r, 1, "TOTAL matched", bold=True, fill=SUBFILL)
    for cc in (2, 3, 6, 7, 8, 9): put(ws, r, cc, "", fill=SUBFILL)
    put(ws, r, 4, f"=SUM(D{first}:D{r-1})", fmt=NUM, bold=True, fill=SUBFILL, align="right")
    put(ws, r, 5, f"=SUM(E{first}:E{r-1})", fmt=NUM, bold=True, fill=SUBFILL, align="right")
    r += 2
    for a, b in FOOTER:
        put(ws, r, 1, a, bold=True, color=NAVY)
        put(ws, r, 2, b)
        for cc in range(3, 10): put(ws, r, cc, "")
        r += 1
    ws.freeze_panes = "C" + str(first)


if __name__ == "__main__":
    import build as base
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    ref = tab_assumptions(wb)
    rows = tab_calllist(wb, ref)
    tab_complexes(wb, ref)
    tab_contacts(wb)
    tab_pf(wb, ref)
    base.tab_mechanics(wb)
    base.tab_script(wb)
    tab_sources(wb)
    order = ["Call List", "PF Targets", "Complexes", "Contacts", "Assumptions",
             "Portaria 140 - Mechanics", "Call Script", "Sources"]
    wb._sheets = [wb[n] for n in order]
    out = "Sutphin_Curtailment_Call_List.xlsx"
    wb.save(out)
    print("wrote", out)
    print("sheets:", wb.sheetnames)
    print(f"named groups on call list: {len(rows[:40])}")
