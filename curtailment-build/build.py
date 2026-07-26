#!/usr/bin/env python3
"""Sutphin — Portaria 140/2026 curtailment call list workbook."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "Sutphin_Curtailment_Call_List.xlsx"

FONT = "Arial"
NAVY = "1F3864"
RULE = "BFBFBF"
HDRFILL = PatternFill("solid", fgColor=NAVY)
SUBFILL = PatternFill("solid", fgColor="D9E2F3")
INPUTFILL = PatternFill("solid", fgColor="FFFF00")
BLUE = "0000FF"          # hardcoded inputs / scenario levers
BLACK = "000000"         # formulas
GREEN = "008000"         # cross-sheet links

THIN = Side(style="thin", color=RULE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def hdr(ws, row, labels, widths=None):
    for i, lab in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=lab)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = HDRFILL
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        c.border = BOX
    ws.row_dimensions[row].height = 30
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")
    ws.row_dimensions[1].height = 20


def put(ws, row, col, value, *, bold=False, size=9, color=BLACK, wrap=True,
        fill=None, fmt=None, italic=False, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, size=size, bold=bold, color=color, italic=italic)
    c.alignment = Alignment(wrap_text=wrap, vertical="top", horizontal=align)
    c.border = BOX
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    return c


# ---------------------------------------------------------------------------
# TAB: Portaria 140 — Mechanics
# ---------------------------------------------------------------------------
MECHANICS = [
    ("THE INSTRUMENT", None, None),
    ("Citation", "Portaria Normativa MME nº 140, de 18 de julho de 2026",
     "DOU 21/07/2026, Ed. 135-A, Seção 1 – Extra A, p. 10. 12 articles + Anexo (minuta do Termo)."),
    ("Statutory basis", "Art. 1º-B, Lei 10.848/2004, inserted by Lei 15.269/2025",
     "Lei 15.269/2025 published DOU 25.11.2025, in force on publication — which is what fixes the window's end date."),
    ("Instrument created", "Termo de Compromisso",
     "Not a termo de adesão, not a transação, not a reconhecimento de dívida."),
    ("Parties", "A UNIÃO, via MME (Poder Concedente) + the GERADOR",
     "CCEE is NOT a party — it is the operational agent. ENBPar is interveniente for Proinfa only."),
    ("Legal nature", "Título executivo extrajudicial (Cl. 11ª, art. 784 II and III, CPC)",
     "Directly enforceable against the União. Forum: Justiça Federal, Seção Judiciária do DF. No arbitration."),
    ("Hierarchy", "Termo > Portaria 140 > CCEE Procedimentos e Regras de Comercialização",
     "Termo, Cl. 5ª, Subcl. 4ª. Matters if CCEE's rules end up narrower than the Termo."),

    ("DATES", None, None),
    ("ELECTION DEADLINE", "10 AUGUST 2026",
     "Art. 2º I — 20 days from publication on 21.07.2026. NOTE: the weekend roll-forward in arts. 4º §4º and 5º §9º does NOT apply to art. 2º. It lands on a Monday, so the point is moot — but assume no elasticity."),
    ("Curtailment window compensated", "1 Sep 2023 to 25 Nov 2025, inclusive",
     "Art. 1º, parágrafo único. End date = entry into force of art. 1º-B."),
    ("CCEE adapted rules due", "~19 Sep 2026", "Art. 7º §1º — 60 days."),
    ("ONS calculation tool + cronograma", "~18 Nov 2026", "Art. 5º §1º I — 120 days."),
    ("ONS final database to CCEE", "~27 Apr 2027", "Chain of art. 5º deadlines."),
    ("Convocação to sign", "Cannot issue before the above", "Art. 2º §3º."),
    ("CCEE publishes cut volumes", "~27 May 2027", "Art. 5º §8º."),
    ("PAYMENT", "Late 2027 base case; H1 2028 on the outer limits",
     "Art. 6º §2º — 180 days from the first business day after the signature deadline. The Portaria fixes NO calendar date, only a chain of relative deadlines."),

    ("WHO IS ELIGIBLE", None, None),
    ("Sources", "Wind and solar FOTOVOLTAICA only",
     "Art. 1º. CSP/heliotérmica is not covered by the wording."),
    ("Connection", "Rede Básica and DITs no âmbito da distribuição",
     "VIRES POINT: the statute says only 'conectada ao SIN'. The Portaria narrows it, arguably excluding SIN-connected plants on non-DIT distribution assets."),
    ("Who holds the right", "Titular de usina com outorga — the grant holder, not the offtaker", ""),
    ("Contract regime", "ACR, ACL and autoprodução are ALL eligible",
     "Termo Cl. 1ª Subcl. 1ª — compensation is due 'independentemente do ambiente ou da modalidade de contratação', with no garantia física cap, no contracted-volume cap, and no frequency or duration threshold."),
    ("Thresholds", "None — no MW floor, no COD cut-off, no minimum curtailment volume", ""),

    ("WHAT IS PAID FOR", None, None),
    ("Indisponibilidade externa", "COMPENSATED — 100%, no haircut", "Termo Cl. 2ª."),
    ("Confiabilidade elétrica", "COMPENSATED — 100%, no haircut", "Termo Cl. 3ª."),
    ("Sobreoferta de energia", "NOT COMPENSATED",
     "Termo Cl. 1ª Subcl. 2ª. This is the majority of curtailed volume and it is the commercial heart of the problem."),
    ("Mixed hours", "Indisponibilidade externa prevails, then confiabilidade; sobreoferta loses",
     "Art. 5º §5º. Generator-favourable tie-break."),
    ("Own-asset carve-out", "Curtailment traced to the generator's own or shared connection assets is NOT compensated",
     "Termo Cl. 2ª/3ª, Subcl. 2ª/3ª."),
    ("Set-off", "Amounts already received under existing rules are deducted",
     "Termo Cl. 5ª Subcl. 1ª VII."),

    ("HOW IT IS VALUED — THE WATERFALL", None, None),
    ("Step 1 — CCEAR-D / CER commitments", "DEBT RELIEF, NOT CASH",
     "Cl. 5ª Subcl. 1ª III: curtailed MWh are treated as energy DELIVERED against CCEAR-D and CER commitments, reducing the generator's delivery-shortfall ressarcimentos. No money moves."),
    ("Step 2 — residual contracted", "Valued on the respective contract terms", "Cl. 5ª Subcl. 1ª IV."),
    ("Step 3 — UNCONTRACTED volume", "CASH, at PLD of the plant's submarket",
     "Cl. 5ª Subcl. 1ª V: 'deverá ser repassada ao GERADOR'. THIS IS THE ONLY TRANCHE THAT PRODUCES A RECEIVABLE ANYONE CAN BUY."),
    ("Step 4 — Proinfa", "At the Proinfa contract price in force at the date of the cut, paid via ENBPar",
     "Cl. 5ª Subcl. 1ª VI; Art. 9º."),
    ("Indexation", "IPCA, from the date of EACH curtailment event to actual payment",
     "Art. 6º §1º. Runs from 2023–2025 event dates, so the accrued uplift by payment is substantial."),
    ("Interest", "SILENT — no juros, no mora, no Selic anywhere in the text",
     "Monetary correction only."),
    ("MCP default-sharing", "CARVED OUT — credits do not participate in the rateio da inadimplência",
     "Art. 6º §6º, unambiguous. The generator does not bear other agents' defaults on this amount."),

    ("WHO ACTUALLY PAYS", None, None),
    ("Funding source", "NOT the CDE, NOT a new encargo",
     "Zero mentions of CDE, encargo or tarifa in the Portaria. Funded under art. 1º-B §5º from unliquidated ressarcimentos OWED BY wind and solar generators under CER and CCEAR-D (modalidade disponibilidade)."),
    ("Mechanism", "Recontabilização financeira de eventos passados, via monthly recontabilisation events",
     "Art. 6º caput and §4º/§5º. Liquidations under ANEEL Despacho 148 of 20.01.2026 are suspended for signatories (§3º) and resumed for all, net of signatories' reduction (§5º)."),
    ("Reported size", "~R$3.3bn of credit against ~R$6bn of generator debt to CCEE",
     "PRESS ESTIMATE ONLY (Poder360). MME's own release states the values are not yet defined and will only be known after the apuração stages."),

    ("THE WAIVER — READ THIS TWICE", None, None),
    ("Scope", "Irrevocable and irretractable waiver of the right to litigate ANY curtailment event to 25 Nov 2025",
     "Termo Cl. 7ª Subcl. 1ª — administrative, arbitral AND judicial, plus discontinuance of any pending action."),
    ("THE ASYMMETRY", "You waive claims on SOBREOFERTA, which the Portaria never pays for",
     "Cl. 1ª Subcl. 2ª excludes sobreoferta from compensation; Cl. 7ª Subcl. 1ª waives claims on all 'eventos de restrição de geração'. This is the single most important commercial point in the whole instrument."),
    ("Collective actions", "Express exclusion from the rol de substituídos",
     "Cl. 7ª Subcl. 3ª — including from the effects of provisional decisions already in force or subsequently granted."),
    ("Injunctions", "Immediate waiver of all suspensive or protective effects of existing court decisions",
     "Cl. 7ª Subcl. 5ª; Cl. 8ª obliges the generator not to oppose provisional orders against the União or CCEE."),
    ("Proof required", "Court petition filed within 10 days of signing; proof to MME",
     "Cl. 7ª Subcl. 4ª. Payment is CONDITIONAL on that proof (Subcl. 7ª)."),
    ("THE ASYMMETRIC SANCTION", "Fail to prove and compensation is suspended while the other obligations survive — including resumption of the ressarcimentos",
     "Cl. 7ª Subcl. 8ª, after a 5-day cure. You can end up having waived, being billed, and not being paid."),
    ("What is NOT waived", "Curtailment events AFTER 25 Nov 2025",
     "The waiver is bounded by event date, not filing date."),
    ("Sucumbência", "None payable by either side", "Cl. 7ª Subcl. 6ª, via art. 1º-B §2º and art. 1º §5º, Lei 9.469/1997."),

    ("TRAPS", None, None),
    ("Garantia física — A BENEFIT, NOT A COST", "Cut volumes count as generation verificada for the annual garantia física revision — 'passiveis ou nao de compensacao'",
     "Art. 10 and Termo Cl. 6a. This PROTECTS the plant: counting curtailed MWh as generated is what stops the physical guarantee being revised DOWN for output the plant was ordered not to produce. It applies to sobreoferta cuts too, which are never paid for in cash — so it is the one place the uncompensated volume still earns something. Methodology still 'a ser definida' by MME, so the value cannot yet be quantified, but the direction is favourable and it belongs on the POSITIVE side of the sign/do-not-sign equation."),
    ("Value it as", "Avoided reduction in future contractable MWh x expected future margin per MWh x duration x methodology probability",
     "Cannot be computed until MME publishes the methodology. Model it as an option, not a number."),
    ("The data trap", "Data submitted under arts. 4º/5º enters the ONS/CCEE databases ONLY if you sign",
     "Art. 11, parágrafo único — a non-signatory's data is 'desconsiderada'. You cannot use the process to improve your dataset and then litigate on it."),
    ("Blind sign-off", "You accept CCEE's published volumes and classifications BEFORE the cash amount is known",
     "Termo Cl. 4ª — declarations are 'irretratável e incondicional', including that the data is 'fidedigna… não tendo sido objeto de qualquer tipo de manipulação'."),
    ("Emolumento", "Payable to CCEE — amount NOT SET in the Portaria",
     "Art. 2º §5º. CCEE will set it."),
    ("ASSIGNABILITY", "THE PORTARIA IS ENTIRELY SILENT",
     "An exhaustive term search returns ZERO occurrences of cessão, cedente/cessionário, transferência, oneração, penhor, caução, garantia real or alienação — in the Portaria or the Termo. Neither authorised nor prohibited. Runs off general law on assignment of credits against the União, plus Cl. 9ª (amendment only by specific instrument between the parties). DO NOT represent the credit as assignable."),

    ("THE FOUR STEPS A GENERATOR MUST TAKE", None, None),
    ("1", "By 10 Aug 2026 — file the manifestação prévia de interesse with MME",
     "Art. 2º I and §1º. Standard form, via the CELEBRA portal at celebra.mme.gov.br."),
    ("2", "Within 10 days of the convocação despacho — file with CCEE (not MME)",
     "Art. 2º II: evidence of legal representation carrying 'poderes específicos e expressos para transigir e renunciar direitos', plus proof of the emolumento."),
    ("3", "Sign the Termo de Compromisso with the Poder Concedente",
     "Art. 2º III — within the period set in the convocação."),
    ("4", "Within 10 days of signing — file the court petition and prove it to MME",
     "Renúncia, desistência, and exclusion from any collective action. Payment is conditional on this."),
]


def tab_mechanics(wb):
    ws = wb.create_sheet("Portaria 140 - Mechanics")
    title(ws, "Portaria Normativa MME nº 140/2026 — what it actually says",
          "Everything below is from the primary text in the DOU of 21 July 2026. Article references are exact. "
          "Where the Portaria is silent, this sheet says so rather than filling the gap.")
    r = 4
    hdr(ws, r, ["Point", "Position", "Detail and article reference"], [34, 46, 88])
    r += 1
    for a, b, c in MECHANICS:
        if b is None:
            cell = put(ws, r, 1, a, bold=True, size=10, color=NAVY, fill=SUBFILL)
            for col in (2, 3):
                put(ws, r, col, "", fill=SUBFILL)
            ws.row_dimensions[r].height = 18
        else:
            put(ws, r, 1, a, bold=True)
            put(ws, r, 2, b)
            put(ws, r, 3, c or "")
        r += 1
    ws.freeze_panes = "A5"
    return ws


# ---------------------------------------------------------------------------
# TAB: Call Script
# ---------------------------------------------------------------------------
from scriptdata import SCRIPT


def tab_script(wb):
    ws = wb.create_sheet("Call Script")
    title(ws, "Call script — the two weeks to 10 August 2026",
          "Advisory first, capital second. The instrument is irrevocable, so the advice is genuinely valuable and "
          "genuinely scarce.")
    r = 4
    hdr(ws, r, ["", "What to say"], [30, 132])
    r += 1
    for a, b in SCRIPT:
        if b is None:
            put(ws, r, 1, a, bold=True, size=10, color=NAVY, fill=SUBFILL)
            put(ws, r, 2, "", fill=SUBFILL)
            ws.row_dimensions[r].height = 18
        else:
            put(ws, r, 1, a, bold=True)
            put(ws, r, 2, b)
        r += 1
    ws.freeze_panes = "A5"
    return ws


if __name__ == "__main__":
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    tab_mechanics(wb)
    tab_script(wb)
    wb.save(OUT)
    print("wrote", OUT, "sheets:", wb.sheetnames)
